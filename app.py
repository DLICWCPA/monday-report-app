import json
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
import io
from flask import Flask, request, jsonify, send_file

# --- HELPER FUNCTIONS FROM YOUR ORIGINAL SCRIPT ---

def format_table(ws, start_row, start_col, num_rows, num_cols, align="center"):
    border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))
    if num_rows == 0: return # Do not format if there are no rows
    for r in range(start_row, start_row + num_rows):
        for c in range(start_col, start_col + num_cols):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r == start_row: cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

def style_last_written_table(ws, title_text, bold_cols=None):
    border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))
    bold_font = Font(bold=True)
    start_header_row = 0
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=1).value == title_text:
            start_header_row = row + 2
            break
    if not start_header_row: return
    num_cols = 0
    while ws.cell(row=start_header_row, column=num_cols + 1).value: num_cols += 1
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=start_header_row, column=col)
        cell.font = bold_font
        cell.border = border
    row = start_header_row + 1
    while ws.cell(row=row, column=1).value is not None and ws.cell(row=row, column=1).value != "":
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if ws.cell(row=row, column=1).value == "Grand Total": cell.font = bold_font
            if bold_cols and ws.cell(row=start_header_row, column=col).value in bold_cols: cell.font = Font(bold=True)
            if col >= 2: cell.alignment = Alignment(horizontal="center")
        row += 1

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ws.merged_cells: continue
            try:
                if cell.value: max_length = max(max_length, len(str(cell.value)))
            except: pass
        adjusted_width = max(max_length + 2, 10)
        ws.column_dimensions[col_letter].width = adjusted_width

def write_merged_title(ws, title, col_span=8, align="center"):
    ws.append([""] * col_span)
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def map_market(country):
    if not isinstance(country, str): return "Others"
    country_upper = country.upper().strip()
    if country_upper == "BRAZIL": return "Brazil Desk"
    if country_upper == "MEXICO": return "Mexico Desk"
    if country_upper == "SPAIN": return "Spain Desk"
    if country_upper == "UNITED KINGDOM": return "UK Desk"
    if country_upper == "AUSTRALIA": return "Australia Desk"
    if country_upper == "INDIA": return "India Desk"
    if country_upper == "UNITED STATES": return "USA Desk"
    if country_upper in ["CHINA", "HONG KONG"]: return "China Desk"
    if country_upper == "GERMANY": return "German Desk"
    if country_upper in ["UNITED ARAB EMIRATES", "UAE", "SAUDI ARABIA", "QATAR", "KUWAIT", "OMAN", "BAHRAIN", "JORDAN", "LEBANON", "ISRAEL", "IRAQ", "IRAN"]: return "Middle East Desk"
    latam_countries = ["ARGENTINA", "COLOMBIA", "PERU", "CHILE", "ECUADOR", "URUGUAY", "PARAGUAY", "BOLIVIA", "COSTA RICA", "PANAMA", "VENEZUELA", "GUATEMALA", "HONDURAS", "EL SALVADOR", "DOMINICAN REPUBLIC", "CUBA", "JAMAICA", "TRINIDAD AND TOBAGO", "BAHAMAS", "BARBADOS", "HAITI", "NICARAGUA", "PUERTO RICO"]
    if country_upper in latam_countries: return "Latam Desk"
    euro_countries = ["FRANCE", "ITALY", "NETHERLANDS", "BELGIUM", "SWEDEN", "NORWAY", "DENMARK", "FINLAND", "SWITZERLAND", "AUSTRIA", "POLAND", "CZECH REPUBLIC", "HUNGARY", "IRELAND", "PORTUGAL", "GREECE", "SLOVAKIA", "SLOVENIA", "ROMANIA", "BULGARIA", "CROATIA", "ESTONIA", "LATVIA", "LITHUANIA", "LUXEMBOURG"]
    if country_upper in euro_countries: return "Euro Desk"
    return "Others"

# --- FLASK APPLICATION SETUP ---
app = Flask(__name__)

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/generate-report', methods=['POST'])
def generate_report():
    try:
        hk_tz = timezone(timedelta(hours=8))
        now_hkt = datetime.now(timezone.utc).astimezone(hk_tz)
        days_until_saturday = (5 - now_hkt.weekday() + 7) % 7
        this_saturday = (now_hkt + timedelta(days=days_until_saturday)).date()
        report_end = this_saturday
        report_begin = report_end - timedelta(days=7)

        data = request.json
        apiKey, boardId = data.get('apiKey'), data.get('boardId')
        apiUrl, headers = "https://api.monday.com/v2", {"Authorization": apiKey}

        all_items, cursor = [], None
        print("Fetching all items from Monday.com...")
        while True:
            query = f"""query {{ boards(ids: {boardId}) {{ items_page(limit:500) {{ cursor items {{ id name column_values {{ text column {{ title }} }} }} }} }} }}"""
            if cursor: query = f"""query {{ next_items_page(cursor: "{cursor}", limit:500) {{ cursor items {{ id name column_values {{ text column {{ title }} }} }} }} }}"""
            
            response = requests.post(apiUrl, json={'query': query}, headers=headers)
            response.raise_for_status()
            monday_data = response.json()
            if "errors" in monday_data: raise Exception(f"Monday.com API Error: {monday_data['errors']}")
            
            page_data = monday_data.get("data", {}).get("next_items_page") or monday_data.get("data", {}).get("boards", [{}])[0].get("items_page", {})
            current_items = page_data.get("items", [])
            if not current_items: break
            all_items.extend(current_items)
            cursor = page_data.get("cursor")
            if not cursor: break
        print(f"✅ Fetched {len(all_items)} total items.")

        rows, columns = [], []
        for item in all_items:
            row_data = {"Item ID": item["id"], "Item Name": item["name"]}
            for col in item.get("column_values", []):
                if col and col.get("column") and col["column"].get("title"):
                    col_title = col["column"]["title"]
                    row_data[col_title] = col["text"]
                    if col_title not in columns: columns.append(col_title)
            rows.append(row_data)
        
        # Ensure essential columns exist before creating the dataframe
        base_cols = ["Item ID", "Item Name"]
        all_cols = base_cols + columns
        df_data = pd.DataFrame(rows, columns=all_cols)

        print("Starting data analysis...")
        df_data["ReportBegin"], df_data["ReportEnd"] = report_begin, report_end
        for col in ["Deal creation date", "Close Date"]:
            if col in df_data.columns: df_data[col] = pd.to_datetime(df_data[col], errors="coerce").dt.date
        
        if 'Group Status' not in df_data.columns: df_data['Group Status'] = "Unknown"
        if 'Potential' not in df_data.columns: df_data['Potential'] = "Unknown"
        if 'Dept' not in df_data.columns: df_data['Dept'] = "Unassigned"
        if 'Country/Region' not in df_data.columns: df_data['Country/Region'] = "Unassigned"
        if 'Referral Source Category' not in df_data.columns: df_data['Referral Source Category'] = "Unknown"

        df_data["IsActiveNow"] = ((((df_data["Deal creation date"] < report_end) & (df_data["Group Status"] == "Active")) | ((df_data["Group Status"] != "Active") & (df_data["Close Date"] >= report_end)))).astype(int)
        df_data["IsActiveBeforeCutoff"] = (((df_data["Deal creation date"] < report_begin) & ((df_data["Group Status"] == "Active") | ((df_data["Group Status"] != "Active") & (df_data["Close Date"] >= report_begin))))).astype(int)
        df_data["AdditionAfterCutoff"] = ((df_data["Deal creation date"] >= report_begin) & (df_data["Deal creation date"] < report_end)).astype(int)
        df_data["RemovalAfterCutoff"] = (df_data["Close Date"].notna() & (df_data["Close Date"] >= report_begin) & (df_data["Close Date"] < report_end)).astype(int)
        df_data["IsHot"] = ((df_data["Potential"] == "Hot") & (df_data["IsActiveNow"] == 1)).astype(int)
        df_data["IsCold"] = ((df_data["Potential"] == "Cold") & (df_data["IsActiveNow"] == 1)).astype(int)

        print("Creating Excel workbook...")
        wb = Workbook()
        wb.remove(wb.active)
        ws_data = wb.create_sheet("Data")
        for r in dataframe_to_rows(df_data, index=False, header=True): ws_data.append(r)
        ws_summary = wb.create_sheet("Summary Report")

        # --- FULL SUMMARY REPORT GENERATION ---
        
        period_from, period_to = report_begin.strftime("%m/%d/%Y"), report_end.strftime("%m/%d/%Y")
        ws_summary.append([""] * 5 + ["Period:", "From", "To"]); [setattr(c, 'font', Font(bold=True)) for c in ws_summary[ws_summary.max_row][-3:]]
        ws_summary.append([""] * 5 + ["", period_from, period_to]); ws_summary.append([])
        
        metric_defs = {"This Week": "IsActiveNow", "Last Week": "IsActiveBeforeCutoff", "Addition (+)": "AdditionAfterCutoff", "Removal (-)": "RemovalAfterCutoff", "Hot": "IsHot", "Cold": "IsCold"}
        h1, v1, h2, v2 = [],[],[],[]
        for k, v in list(metric_defs.items())[:4]: h1.append(k); v1.append(df_data[v].sum())
        for k, v in list(metric_defs.items())[4:]: h2.append(k); v2.append(df_data[v].sum())
        ws_summary.append([""]*1+["Enquiries Movement"]); ws_summary.cell(row=ws_summary.max_row, column=2).font=Font(bold=True); ws_summary.append([]); ws_summary.append([""]*1+h1); s=ws_summary.max_row; ws_summary.append([""]*1+v1); format_table(ws_summary,s,2,2,len(h1)); ws_summary.append([])
        ws_summary.append([""]*1+["Enquiries by Potential"]); ws_summary.cell(row=ws_summary.max_row, column=2).font=Font(bold=True); ws_summary.append([]); ws_summary.append([""]*1+["Potential"]+h2); s=ws_summary.max_row; ws_summary.append([""]*1+[""]+v2); format_table(ws_summary,s,2,2,1+len(h2)); ws_summary.append([])

        for t in ["Addition", "Removal"]:
            df_s = df_data[df_data[f"{t}AfterCutoff"] == 1]
            cols = ["Dept", "Item Name", "Country/Region", "Salesperson", "Service", "Stage", "Referral Source Category", "Group Status"]
            title = f"Enquiries {t.replace('ition','ed').replace('al','ed')} This Week"
            write_merged_title(ws_summary, title, align="left"); ws_summary.append([])
            ws_summary.append(cols); s = ws_summary.max_row
            if not df_s.empty:
                for _, r in df_s[cols].iterrows(): ws_summary.append(list(r))
            format_table(ws_summary, s, 1, 1+len(df_s), len(cols), align="left"); ws_summary.append([])

        df_active = df_data[df_data["IsActiveNow"] == 1].copy()
        if not df_active.empty:
            df_active.columns = df_active.columns.str.strip()
            
            # Country Matrix
            matrix = pd.pivot_table(df_active,index="Country/Region",columns="Potential",aggfunc="size",fill_value=0)
            if not matrix.empty:
                matrix["Total"] = matrix.sum(axis=1)
                matrix_sorted = matrix.sort_values(by="Total", ascending=False)
                total_row = matrix_sorted.sum(numeric_only=True).to_frame().T; total_row.index = ["Grand Total"]
                final_matrix = pd.concat([matrix_sorted, total_row]).reset_index()
                write_merged_title(ws_summary, "Active Enquiries by Country and Potential", align="left"); ws_summary.append([])
                for r in dataframe_to_rows(final_matrix, index=False, header=True): ws_summary.append(r)
                style_last_written_table(ws_summary, "Active Enquiries by Country and Potential", bold_cols=["Total"]); ws_summary.append([])

            # Market Division Matrix
            df_active["Market Division"] = df_active["Country/Region"].apply(map_market)
            pivot = pd.pivot_table(df_active,index="Market Division",columns="Potential",aggfunc="size",fill_value=0)
            if not pivot.empty:
                pivot["Total"] = pivot.sum(axis=1)
                pivot_sorted = pivot.sort_values(by="Total", ascending=False)
                total_row = pivot_sorted.sum(numeric_only=True).to_frame().T; total_row.index = ["Grand Total"]
                final_matrix = pd.concat([pivot_sorted, total_row]).reset_index().rename(columns={"index": "Market Segment"})
                write_merged_title(ws_summary, "Active Enquiries by Market Division and Potential (7+4 Desk Mapping)", align="left"); ws_summary.append([])
                for r in dataframe_to_rows(final_matrix, index=False, header=True): ws_summary.append(r)
                style_last_written_table(ws_summary, "Active Enquiries by Market Division and Potential (7+4 Desk Mapping)", bold_cols=["Total"]); ws_summary.append([])

        # Referral Source
        total_by_source = df_data.groupby("Referral Source Category").size().rename("Total")
        won_by_source = df_data[df_data["Group Status"] == "Won"].groupby("Referral Source Category").size().rename("Won")
        effectiveness = pd.concat([total_by_source, won_by_source], axis=1).fillna(0).astype(int)
        if not effectiveness.empty:
            effectiveness["Win %"] = ((effectiveness["Won"] / effectiveness["Total"].replace(0,1)) * 100).round(1).astype(str) + "%"
            effectiveness = effectiveness.reset_index()
            grand_total = {"Referral Source Category": "Grand Total", "Total": effectiveness["Total"].sum(), "Won": effectiveness["Won"].sum()}
            grand_total["Win %"] = str(round((grand_total["Won"] / grand_total["Total"] if grand_total["Total"] > 0 else 0) * 100, 1)) + "%"
            effectiveness = pd.concat([effectiveness, pd.DataFrame([grand_total])], ignore_index=True)
            write_merged_title(ws_summary, "Referral Source Effectiveness (Based on 'Won' Deals)", align="left"); ws_summary.append([])
            for r in dataframe_to_rows(effectiveness, index=False, header=True): ws_summary.append(r)
            style_last_written_table(ws_summary, "Referral Source Effectiveness (Based on 'Won' Deals)", bold_cols=["Win %"]); ws_summary.append([])

        # --- Department and Desk Breakdowns ---
        section_cols = ["Item Name", "Country/Region", "Salesperson", "Service", "Stage", "Potential", "Referral Source Category"]
        
        # Departments
        ws_summary.append([]); write_merged_title(ws_summary, "Breakdown by Departments", col_span=len(section_cols)); ws_summary.append([])
        depts_to_report = ["COS", "CCT-GBA", "CCT-SH", "AG2", "TAX"]
        for dept in depts_to_report:
            df_dept = df_active[df_active["Dept"].fillna("").str.contains(dept, case=False)]
            write_merged_title(ws_summary, dept, col_span=len(section_cols))
            summary = f"There are total {len(df_dept)} active enquiries for {dept}, out of which {df_dept['IsHot'].sum()} are hot and {df_dept['IsCold'].sum()} are cold."
            write_merged_title(ws_summary, summary, col_span=len(section_cols), align="left"); ws_summary.append([])
            ws_summary.append(section_cols); s = ws_summary.max_row
            if not df_dept.empty:
                for _, r in df_dept[section_cols].iterrows(): ws_summary.append(list(r))
            format_table(ws_summary, s, 1, 1+len(df_dept), len(section_cols), align="left"); ws_summary.append([])

        # Desks
        ws_summary.append([]); write_merged_title(ws_summary, "Individual Desks", col_span=len(section_cols)); ws_summary.append([])
        df_active["Market Division"] = df_active["Country/Region"].apply(map_market)
        desks_to_report = [d for d in df_active["Market Division"].unique() if d != "Others"]
        for desk in desks_to_report:
            df_desk = df_active[df_active["Market Division"] == desk]
            write_merged_title(ws_summary, desk, col_span=len(section_cols))
            summary = f"There are total {len(df_desk)} active enquiries for {desk}, out of which {df_desk['IsHot'].sum()} are hot and {df_desk['IsCold'].sum()} are cold."
            write_merged_title(ws_summary, summary, col_span=len(section_cols), align="left"); ws_summary.append([])
            ws_summary.append(section_cols); s = ws_summary.max_row
            if not df_desk.empty:
                for _, r in df_desk[section_cols].iterrows(): ws_summary.append(list(r))
            format_table(ws_summary, s, 1, 1+len(df_desk), len(section_cols), align="left"); ws_summary.append([])
            
        # Salesperson
        ws_summary.append([]); write_merged_title(ws_summary, "Breakdown by Salesperson", col_span=5); ws_summary.append([])
        sales_cols = ["Item Name", "Dept", "Service", "Potential"]
        ws_summary.append(["Salesperson"] + sales_cols); s = ws_summary.max_row
        df_sales = df_active[df_active["Salesperson"].notna()]
        if not df_sales.empty:
            for salesperson, group in df_sales.groupby("Salesperson"):
                ws_summary.append([f"{salesperson} (Total: {len(group)})"] + [""] * (len(sales_cols)))
                for _, r in group.iterrows():
                    ws_summary.append([""] + [r[c] for c in sales_cols])
        format_table(ws_summary, s, 1, ws_summary.max_row - s + 1, 1+len(sales_cols), align="left")

        print("Formatting and saving file...")
        autofit_columns(ws_summary)
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        filename = f"Weekly_Enquiries_Report_{report_begin}_to_{report_end}.xlsx"
        return send_file(file_stream, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500

# --- RUN THE FLASK APPLICATION ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)