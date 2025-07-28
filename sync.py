import json
import requests
import pandas as pd
import os
from datetime import datetime, timedelta, timezone, date
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

# --- Helper Functions

def format_table(ws, start_row, start_col, num_rows, num_cols, align="center"):
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )

    for r in range(start_row, start_row + num_rows):
        for c in range(start_col, start_col + num_cols):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if r == start_row:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)


def style_last_written_table(ws, title_text, bold_cols=None):
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    bold_font = Font(bold=True)

    start_header_row = -1
    for row in range(ws.max_row, 0, -1):
        if ws.cell(row=row, column=1).value == title_text:
            start_header_row = row + 2
            break
    if start_header_row == -1:
        return

    num_cols = 0
    while ws.cell(row=start_header_row, column=num_cols + 1).value:
        num_cols += 1

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=start_header_row, column=col)
        cell.font = bold_font
        cell.border = border

    row = start_header_row + 1
    while ws.cell(row=row, column=1).value is not None:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border

            if ws.cell(row=row, column=1).value == "Grand Total":
                cell.font = bold_font

            if bold_cols and ws.cell(row=start_header_row, column=col).value in bold_cols:
                cell.font = bold_font

            if col >= 2:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        row += 1

def write_merged_title(ws, title, col_span=8, align="center"):
    ws.append([""] * col_span)
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)

    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def map_market(country):
    if country == "Brazil": return "Brazil Desk"
    if country == "Mexico": return "Mexico Desk"
    if country == "Spain": return "Spain Desk"
    if country == "United Kingdom": return "UK Desk"
    if country == "Australia": return "Australia Desk"
    if country == "India": return "India Desk"
    if country == "United States": return "USA Desk"
    if country in ["China", "Hong Kong"]: return "China Desk"
    if country == "Germany": return "German Desk"
    if country in ["United Arab Emirates", "Saudi Arabia", "Qatar", "Kuwait", "Oman", "Bahrain", "Jordan", "Lebanon", "Israel"]: return "Middle East Desk"
    if country in ["Argentina", "Colombia", "Peru", "Chile", "Ecuador", "Uruguay", "Paraguay", "Bolivia", "Costa Rica", "Panama", "Venezuela", "Guatemala", "Honduras", "El Salvador", "Dominican Republic", "Cuba", "Jamaica", "Trinidad and Tobago", "Bahamas", "Barbados", "Haiti", "Nicaragua"]: return "Latam Desk"
    if country in ["France", "Italy", "Netherlands", "Belgium", "Sweden", "Norway", "Denmark", "Finland", "Switzerland", "Austria", "Poland", "Czech Republic", "Hungary", "Ireland", "Portugal", "Greece", "Slovakia", "Slovenia", "Romania", "Bulgaria", "Croatia", "Estonia", "Latvia", "Lithuania", "Luxembourg"]: return "Euro Desk"
    else: return "Others"

# --- Main Report Generation Function ---

def generate_report(report_begin: date, report_end: date):
    api_key = os.environ.get("MONDAY_API_KEY") 
    board_id = 3678769221
    if not api_key:
        raise ValueError("MONDAY_API_KEY environment variable not set.")

    def fetch_items(cursor=None):
        query = f"""query {{ boards(ids: {board_id}) {{ items_page {{ cursor items {{ id name column_values {{ text column {{ title }} }} }} }} }} }}"""
        if cursor:
            query = f"""query {{ next_items_page(cursor: "{cursor}") {{ cursor items {{ id name column_values {{ text column {{ title }} }} }} }} }}"""
        response = requests.post("https://api.monday.com/v2", json={"query": query}, headers={"Authorization": api_key})
        return response.json() if response.status_code == 200 else None

    all_items = []
    cursor = None
    while True:
        data = fetch_items(cursor)
        if not data or "data" not in data or not data["data"]: break
        page_data = data["data"].get("next_items_page") or (data["data"]["boards"][0]["items_page"] if data["data"]["boards"] else None)
        if not page_data: break
        all_items.extend(page_data.get("items", []))
        cursor = page_data.get("cursor")
        if not cursor: break
    
    rows = []
    for item in all_items:
        row = {"Item ID": item["id"], "Item Name": item["name"]}
        for col in item["column_values"]:
            if col.get("column") and "title" in col["column"]:
                row[col["column"]["title"]] = col["text"]
        rows.append(row)
    
    if not rows:
        df_data = pd.DataFrame()
    else:
        df_data = pd.DataFrame(rows)
    
    df_data["ReportBegin"] = report_begin
    df_data["ReportEnd"] = report_end
    df_data["Deal creation date"] = pd.to_datetime(df_data["Deal creation date"], errors="coerce").dt.date
    df_data["Close Date"] = pd.to_datetime(df_data["Close Date"], errors="coerce").dt.date
    df_data["IsActiveBeforeCutoff"] = (((df_data["Deal creation date"] < df_data["ReportBegin"]) & ((df_data["Group Status"] == "Active") | ((df_data["Group Status"] != "Active") & (df_data["Close Date"] >= df_data["ReportBegin"]))))).astype(int)
    df_data["IsActiveNow"] = ((((df_data["Deal creation date"] < df_data["ReportEnd"]) & (df_data["Group Status"] == "Active")) | ((df_data["Group Status"] != "Active") & (df_data["Close Date"] >= df_data["ReportEnd"])))).astype(int)
    df_data["AdditionAfterCutoff"] = ((df_data["Deal creation date"] >= df_data["ReportBegin"]) & (df_data["Deal creation date"] < df_data["ReportEnd"])).astype(int)
    df_data["RemovalAfterCutoff"] = (df_data["Close Date"].notna() & (df_data["Close Date"] >= df_data["ReportBegin"]) & (df_data["Close Date"] < df_data["ReportEnd"])).astype(int)
    df_data["IsHot"] = ((df_data["Potential"] == "Hot") & (df_data["IsActiveNow"] == 1)).astype(int)
    df_data["IsCold"] = ((df_data["Potential"] == "Cold") & (df_data["IsActiveNow"] == 1)).astype(int)

    excel_buffer = BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df_data.to_excel(writer, sheet_name="Data", index=False)
        
        wb = writer.book
        if "Summary Report" in wb.sheetnames:
            del wb["Summary Report"]
        ws_summary = wb.create_sheet("Summary Report")

    
        # ----------------------------------------
        # 📅 Reporting Period Table 
        # ----------------------------------------

        
        # Format the dates in MM/DD/YYYY format
        period_from = df_data["ReportBegin"].iloc[0].strftime("%m/%d/%Y")
        period_to = df_data["ReportEnd"].iloc[0].strftime("%m/%d/%Y")

        # Row 1: Period labels
        ws_summary.append(["Period:", "From", "To"])
        ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
        ws_summary.cell(row=ws_summary.max_row, column=2).font = Font(bold=True)
        ws_summary.cell(row=ws_summary.max_row, column=3).font = Font(bold=True)

        # Row 2: Period values
        ws_summary.append(["", period_from, period_to])

        # ----------------------------------------
        # ➖ Horizontal Line (separator across A to H)
        # ----------------------------------------

        # Insert a horizontal line row (one row of empty cells with bottom borders)
        line_row = ws_summary.max_row + 1
        ws_summary.append([""] * 8)  # 8 columns: A to H

        # Apply bottom border to the separator row
        thin_border = Border(bottom=Side(style="thin", color="000000"))
        for col in range(1, 9):  # Columns A to H
            cell = ws_summary.cell(row=line_row, column=col)
            cell.border = thin_border

        # Add a spacer row after the line
        ws_summary.append([])

        # ----------------------------------------
        # 📊 Enquiries Movement Table 
        # ----------------------------------------

        # Define metric labels and column mappings
        metric_definitions = {
            "This Week": "IsActiveNow",
            "Last Week": "IsActiveBeforeCutoff",
            "Addition (+)": "AdditionAfterCutoff",
            "Removal (-)": "RemovalAfterCutoff",
            "Hot": "IsHot",
            "Cold": "IsCold"
        }

        # Build summary values
        row1_headers = []
        row1_values = []
        row2_headers = []
        row2_values = []

        for label, col in list(metric_definitions.items())[:4]:  # Movement metrics
            if col in df_data.columns:
                row1_headers.append(label)
                row1_values.append(df_data[col].sum())

        for label, col in list(metric_definitions.items())[4:]:  # Status breakdown
            if col in df_data.columns:
                row2_headers.append(label)
                row2_values.append(df_data[col].sum())

        # Title row in column BB
        ws_summary.append([""] * 1 + ["Enquiries Movement"])
        ws_summary.cell(row=ws_summary.max_row, column=2).font = Font(bold=True)

        # Spacer row (left empty)
        ws_summary.append([])

        # Write header row
        ws_summary.append([""] * 1 + row1_headers)

        # Now correctly capture the start of the table
        start_row = ws_summary.max_row

        # Write data row
        ws_summary.append([""] * 1 + row1_values)


        # Apply formatting
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=2,  # Column B
            num_rows=2,
            num_cols=len(row1_headers)  # Include the blank offset columns
        )

        # Spacer
        ws_summary.append([])

        # ----------------------------------------
        # 📊 Enquiries by Potential Table (aligned with Enquiries Movement)
        # ----------------------------------------

        # Title row starting in Column B
        ws_summary.append([""] * 1 + ["Enquiries by Potential"])
        ws_summary.cell(row=ws_summary.max_row, column=2).font = Font(bold=True)


        # Blank spacer row
        ws_summary.append([])

        # Write header row (with left padding)
        ws_summary.append([""] * 1 + ["Potential"] + row2_headers)

        # Capture header row for styling
        start_row = ws_summary.max_row

        # Write data row (with left padding)
        ws_summary.append([""] * 1 + [""] + row2_values)

        # Apply formatting
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=2,  # Column B
            num_rows=2,
            num_cols=1 + len(row2_headers)  # 1 for "Potential" label + data columns
        )

        ws_summary.append([""] * 8)  # Creates an empty row with 8 blank cells

        # ----------------------------------------
        # ➕ Enquiries Added This Week Table (styled, starts at Column A, left-aligned)
        # ----------------------------------------

        # Filter rows where AdditionAfterCutoff == 1
        added_df = df_data[df_data["AdditionAfterCutoff"] == 1]

        # Define columns to show
        added_columns = [
            "Dept", "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Referral Source Category", "Group Status"
        ]

        # Insert title at column A
        write_merged_title(ws_summary, "Enquiries Added This Week", align="left")

        # Spacer row
        ws_summary.append([])

        # Write header row
        ws_summary.append(added_columns)
        start_row = ws_summary.max_row  # Capture header row

        # Write each data row
        for _, row in added_df[added_columns].iterrows():
            ws_summary.append(list(row))

        # Apply formatting with left alignment
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,  # Column A
            num_rows=1 + len(added_df),  # header + data
            num_cols=len(added_columns),
            align="left"  # 👈 Make everything left-aligned
        )

        # ----------------------------------------
        # ➖ Enquiries Removed This Week Table (styled, starts at Column A, left-aligned)
        # ----------------------------------------

        # Add spacer row before section
        ws_summary.append([])

        # Filter rows where RemovalAfterCutoff == 1
        removed_df = df_data[df_data["RemovalAfterCutoff"] == 1]

        # Define columns to show
        removed_columns = [
            "Dept", "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Referral Source Category", "Group Status"
        ]

        # Insert title
        write_merged_title(ws_summary, "Enquiries Removed This Week", align="left")


        # Spacer row
        ws_summary.append([])

        # Write header row
        ws_summary.append(removed_columns)
        start_row = ws_summary.max_row

        # Write data rows
        for _, row in removed_df[removed_columns].iterrows():
            ws_summary.append(list(row))

        # Apply formatting
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,  # Column A
            num_rows=1 + len(removed_df),
            num_cols=len(removed_columns),
            align="left"
        )

        # Add spacer row after the section (optional, for next table)
        ws_summary.append([])

        # --------------------------------------------
        # 📊 Add Matrix: Active Enquiries by Country and Potential (Sorted + Total + Grand Total Row)
        # --------------------------------------------

        if "IsActiveNow" in df_data.columns:
            df_active = df_data[df_data["IsActiveNow"] == 1]
            if "Country/Region" in df_active.columns and "Potential" in df_active.columns:
                matrix = pd.pivot_table(df_active, index="Country/Region", columns="Potential", aggfunc="size", fill_value=0)
                matrix["Total"] = matrix.sum(axis=1)
                matrix_sorted = matrix.sort_values(by="Total", ascending=False)
                total_row = matrix_sorted.sum(numeric_only=True).to_frame().T
                total_row.index = ["Grand Total"]
                final_matrix = pd.concat([matrix_sorted, total_row]).reset_index()
                
                ws_summary.append([])
                write_merged_title(ws_summary, "Active Enquiries by Country and Potential", align="left")
                ws_summary.append([])
                for row in dataframe_to_rows(final_matrix, index=False, header=True):
                    ws_summary.append(row)
                style_last_written_table(ws_summary, "Active Enquiries by Country and Potential", bold_cols=["Total"])


        # --------------------------------------------
        # 📊 Add Matrix: Active Enquiries by 7+4 Market Division and Potential
        # --------------------------------------------


        if "IsActiveNow" in df_data.columns and "Country/Region" in df_data.columns:
            df_active = df_data[df_data["IsActiveNow"] == 1].copy()
            df_active["Market Division"] = df_active["Country/Region"].map(map_market)

            # 3. Create pivot table: Market Division × Potential
            pivot = pd.pivot_table(
                df_active,
                index="Market Division",
                columns="Potential",
                aggfunc="size",
                fill_value=0
            )

            pivot["Total"] = pivot.sum(axis=1)
            pivot_sorted = pivot.sort_values(by="Total", ascending=False)

            # ✅ Set index name on both DataFrames
            pivot_sorted.index.name = "Market Segment"
            total_row = pivot_sorted.sum(numeric_only=True).to_frame().T
            total_row.index = ["Grand Total"]
            total_row.index.name = "Market Segment"

            # Combine and reset
            final_matrix = pd.concat([pivot_sorted, total_row]).reset_index()



            # Add styled matrix: Market Division vs Potential
            ws_summary.append([])
            
            write_merged_title(ws_summary, "Active Enquiries by Market Division and Potential (7+4 Desk Mapping)", align="left")

            ws_summary.append([])

            for row in dataframe_to_rows(final_matrix, index=False, header=True):
                ws_summary.append(row)

            style_last_written_table(
                ws_summary,
                "Active Enquiries by Market Division and Potential (7+4 Desk Mapping)",
                bold_cols=["Total"]
            )


        # --------------------------------------------
        # 🎯 Referral Source Effectiveness Based on Wins (With % and Grand Total)
        # --------------------------------------------

    

        # Proceed only if necessary columns exist
        if "Referral Source Category" in df_data.columns and "Group Status" in df_data.columns:
            # Count total enquiries per source
            total_by_source = df_data.groupby("Referral Source Category").size().rename("Total")

            # Count 'Won' enquiries per source
            won_by_source = df_data[df_data["Group Status"] == "Won"] \
                .groupby("Referral Source Category").size().rename("Won")

            # Combine into a single DataFrame
            effectiveness = pd.concat([total_by_source, won_by_source], axis=1).fillna(0)

            # Calculate Win % and format as string with %
            effectiveness["Win %"] = (
                (effectiveness["Won"] / effectiveness["Total"]) * 100
            ).round(1).astype(str) + "%"

            # Reset index and sort by Win %
            effectiveness = effectiveness.reset_index()
            effectiveness["Win % (sort)"] = effectiveness["Won"] / effectiveness["Total"]
            effectiveness = effectiveness.sort_values(by="Win % (sort)", ascending=False).drop(columns=["Win % (sort)"])

            # Add Grand Total row
            grand_total = {
                "Referral Source Category": "Grand Total",
                "Total": int(effectiveness["Total"].sum()),
                "Won": int(effectiveness["Won"].sum()),
                "Win %": str(round(effectiveness["Won"].sum() / effectiveness["Total"].sum() * 100, 1)) + "%"
            }
            effectiveness.loc[len(effectiveness)] = grand_total

            # Append table to Summary Report
            ws_summary.append([])
            write_merged_title(ws_summary, "Referral Source Effectiveness (Based on 'Won' Deals)", align="left")

            # Add an empty row for spacing
            ws_summary.append([])


            for row in dataframe_to_rows(effectiveness, index=False, header=True):
                ws_summary.append(row)

        
        # --------------------------------------------
        # 🎨 Style the Referral Source Effectiveness Table
        # --------------------------------------------
        
        # Find the header row for the effectiveness table
        last_row = ws_summary.max_row
        for row in range(last_row, 0, -1):
            if ws_summary.cell(row=row, column=1).value == "Referral Source Effectiveness (Based on 'Won' Deals)":
                # Skip the blank row we added — actual table header is 2 rows after title
                start_row = row + 2
                break


        # Count how many columns exist in the header
        num_cols = 0
        while ws_summary.cell(row=start_row, column=num_cols + 1).value:
            num_cols += 1

        # Define bold font and black border
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # Apply styling to header row
        for col in range(1, num_cols + 1):
            cell = ws_summary.cell(row=start_row, column=col)
            cell.font = bold_font
            cell.border = border

    

        # Style the rest of the table
        row = start_row + 1
        while ws_summary.cell(row=row, column=1).value:
            for col in range(1, num_cols + 1):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border

                # Center-align columns from 2nd onward
                if col >= 2:
                    cell.alignment = Alignment(horizontal="center")

                # Bold the Grand Total row
                if ws_summary.cell(row=row, column=1).value == "Grand Total":
                    cell.font = bold_font

                # Bold the Win % column
                header = ws_summary.cell(row=start_row, column=col).value
                if header == "Win %":
                    cell.font = bold_font
            row += 1


        # ----------------------------------------
        # 📊 Section: Breakdown by Departments (styled title row)
        # ----------------------------------------

        # Add spacer before the section
        ws_summary.append([""] * 8)  # Creates an empty row with 8 blank cells

        # Create and merge the title row
        title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)

        cell = ws_summary.cell(row=title_row, column=1)
        cell.value = "Breakdown by Departments"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Spacer before any new block
        ws_summary.append([""] * 8)  # Enforced row A–H


        # ----------------------------------------
        # 🔹 Subsection: COS (merged, center-aligned, plain style)
        # ----------------------------------------

        # Add spacer row
        ws_summary.append([])

        # Add "COS" merged row across A–H
        cos_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=cos_row, start_column=1, end_row=cos_row, end_column=8)

        cell = ws_summary.cell(row=cos_row, column=1)
        cell.value = "COS"
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # ✅ DEFINE cos_df HERE so it's ready
        cos_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Dept"].fillna("").str.upper() == "COS")
        ]
        # Spacer before any new block
        ws_summary.append([""] * 8)  # Enforced row A–H

        # ----------------------------------------
        # 🔸 COS summary line with red bold numbers (fallback: bold only)
        # ----------------------------------------

        # Add spacer
        ws_summary.append([])

        # Count
        total_cos = int(len(cos_df))
        hot_cos = int(cos_df["IsHot"].sum()) if "IsHot" in cos_df.columns else 0
        cold_cos = int(cos_df["IsCold"].sum()) if "IsCold" in cos_df.columns else 0

        # Build fallback summary string
        summary_text = f"There are total {total_cos} active enquiries for COS, out of which {hot_cos} are hot and {cold_cos} are cold."

        # Insert merged row
        summary_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True, color="000000")  # All bold, black font
        cell.alignment = Alignment(horizontal="left", vertical="center")



        # ----------------------------------------
        # 📋 COS Enquiries Table (no index, styled like Enquiries Added, starts at A, left-aligned)
        # ----------------------------------------

        # Add spacer row before table
        ws_summary.append([])

        # Filter data for Active COS enquiries
        cos_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Dept"].fillna("").str.upper() == "COS")
        ]

        # Define columns for display
        cos_columns = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        # Insert header row
        ws_summary.append(cos_columns)
        start_row = ws_summary.max_row

        # Write data rows
        for _, row in cos_df[cos_columns].iterrows():
            ws_summary.append(list(row))

        # Apply formatting
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,  # Column A
            num_rows=1 + len(cos_df),
            num_cols=len(cos_columns),
            align="left"
        )

        # ----------------------------------------
        # 🟦 CCT-GBA Section
        # ----------------------------------------

        # Add visual spacer row
        ws_summary.append([""] * 8)

        # Insert merged title row: CCT-GBA
        cct_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=cct_title_row, start_column=1, end_row=cct_title_row, end_column=8)

        cct_title_cell = ws_summary.cell(row=cct_title_row, column=1)
        cct_title_cell.value = "CCT-GBA"
        cct_title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Filter Active Now enquiries for CCT-GBA
        cct_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Dept"].fillna("").str.upper() == "CCT-GBA")
        ]

        # Add summary line: total / hot / cold
        total_cct = len(cct_df)
        hot_cct = cct_df["IsHot"].sum() if "IsHot" in cct_df.columns else 0
        cold_cct = cct_df["IsCold"].sum() if "IsCold" in cct_df.columns else 0

        ws_summary.append([""] * 8)  # spacer before summary
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total_cct} active enquiries for CCT-GBA, out of which {hot_cct} are hot and {cold_cct} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Insert enquiry table
        ws_summary.append([])

        cct_columns = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        # Header row
        ws_summary.append(cct_columns)
        start_row = ws_summary.max_row

        # Data rows
        for _, row in cct_df[cct_columns].iterrows():
            ws_summary.append(list(row))

        # Style the table
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(cct_df),
            num_cols=len(cct_columns),
            align="left"
        )

        # ----------------------------------------
        # 🟦 CCT-SH Section
        # ----------------------------------------

        # Add spacer row
        ws_summary.append([""] * 8)

        # Merged title row: CCT-SH
        cctsh_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=cctsh_title_row, start_column=1, end_row=cctsh_title_row, end_column=8)

        cctsh_title_cell = ws_summary.cell(row=cctsh_title_row, column=1)
        cctsh_title_cell.value = "CCT-SH"
        cctsh_title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Filter data for Active Now enquiries in CCT-SH
        cctsh_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Dept"].fillna("").str.upper() == "CCT-SH")
        ]

        # Add summary
        total_cctsh = len(cctsh_df)
        hot_cctsh = cctsh_df["IsHot"].sum() if "IsHot" in cctsh_df.columns else 0
        cold_cctsh = cctsh_df["IsCold"].sum() if "IsCold" in cctsh_df.columns else 0

        ws_summary.append([""] * 8)  # spacer before summary
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total_cctsh} active enquiries for CCT-SH, out of which {hot_cctsh} are hot and {cold_cctsh} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Add enquiry table
        ws_summary.append([])

        cctsh_columns = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cctsh_columns)
        start_row = ws_summary.max_row

        for _, row in cctsh_df[cctsh_columns].iterrows():
            ws_summary.append(list(row))

        # Apply styling
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(cctsh_df),
            num_cols=len(cctsh_columns),
            align="left"
        )

        # ----------------------------------------
        # 🟦 AG2 Section
        # ----------------------------------------

        # Add spacer row
        ws_summary.append([""] * 8)

        # Merged title row: AG2
        ag2_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=ag2_title_row, start_column=1, end_row=ag2_title_row, end_column=8)

        ag2_title_cell = ws_summary.cell(row=ag2_title_row, column=1)
        ag2_title_cell.value = "AG2"
        ag2_title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Filter data for Active Now enquiries in AG2
        ag2_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Dept"].fillna("").str.contains("AG2", case=False))
        ]


        # Add summary
        total_ag2 = len(ag2_df)
        hot_ag2 = ag2_df["IsHot"].sum() if "IsHot" in ag2_df.columns else 0
        cold_ag2 = ag2_df["IsCold"].sum() if "IsCold" in ag2_df.columns else 0

        ws_summary.append([""] * 8)  # spacer before summary
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total_ag2} active enquiries for AG2, out of which {hot_ag2} are hot and {cold_ag2} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Add enquiry table
        ws_summary.append([])

        ag2_columns = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(ag2_columns)
        start_row = ws_summary.max_row

        for _, row in ag2_df[ag2_columns].iterrows():
            ws_summary.append(list(row))

        # Apply table formatting
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(ag2_df),
            num_cols=len(ag2_columns),
            align="left"
        )

        # ----------------------------------------
        # 🟦 TAX Section
        # ----------------------------------------

        # Add spacer row
        ws_summary.append([""] * 8)

        # Merged title row: TAX
        tax_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=tax_title_row, start_column=1, end_row=tax_title_row, end_column=8)

        tax_title_cell = ws_summary.cell(row=tax_title_row, column=1)
        tax_title_cell.value = "TAX"
        tax_title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Filter Active Now enquiries for departments that CONTAIN "TAX"
        tax_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Dept"].fillna("").str.contains("TAX", case=False))
        ]

        # Summary values
        total_tax = len(tax_df)
        hot_tax = tax_df["IsHot"].sum() if "IsHot" in tax_df.columns else 0
        cold_tax = tax_df["IsCold"].sum() if "IsCold" in tax_df.columns else 0

        # Add summary
        ws_summary.append([""] * 8)  # spacer before summary
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total_tax} active enquiries for TAX, out of which {hot_tax} are hot and {cold_tax} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Add enquiry table
        ws_summary.append([])

        tax_columns = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(tax_columns)
        start_row = ws_summary.max_row

        for _, row in tax_df[tax_columns].iterrows():
            ws_summary.append(list(row))

        # Apply formatting
        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(tax_df),
            num_cols=len(tax_columns),
            align="left"
        )

        # ----------------------------------------
        # 📊 Section: Individual Desks (styled section header like Breakdown by Departments)
        # ----------------------------------------

        # Add spacer row before section
        ws_summary.append([""] * 8)

        # Define the section title row and merge range
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        # Set title cell
        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = "Individual Desks"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Add empty row after the section title
        ws_summary.append([""] * 8)

        # ----------------------------------------
        # 🟩 Market Segment: Brazil Desk
        # ----------------------------------------

        market_segment_name = "Brazil Desk"

        # Add spacer row
        ws_summary.append([""] * 8)

        # Insert merged title row
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Filter enquiries for Brazil (Active only)
        segment_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "BRAZIL")
        ]

        # Add summary
        total_seg = len(segment_df)
        hot_seg = segment_df["IsHot"].sum() if "IsHot" in segment_df.columns else 0
        cold_seg = segment_df["IsCold"].sum() if "IsCold" in segment_df.columns else 0

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total_seg} active enquiries for {market_segment_name}, out of which {hot_seg} are hot and {cold_seg} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in segment_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(segment_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: Mexico Desk
        # ----------------------------------------

        market_segment_name = "Mexico Desk"

        # Add spacer row
        ws_summary.append([""] * 8)

        # Insert merged title row
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Filter enquiries for Mexico (Active only)
        segment_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "MEXICO")
        ]

        # Add summary
        total_seg = len(segment_df)
        hot_seg = segment_df["IsHot"].sum() if "IsHot" in segment_df.columns else 0
        cold_seg = segment_df["IsCold"].sum() if "IsCold" in segment_df.columns else 0

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total_seg} active enquiries for {market_segment_name}, out of which {hot_seg} are hot and {cold_seg} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in segment_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(segment_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: Latam Desk
        # ----------------------------------------

        market_segment_name = "Latam Desk"

        # Define exclusion list (Mexico, Brazil)
        excluded = ["MEXICO", "BRAZIL"]

        # Latin American & Caribbean countries - approximate logic
        latam_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].notna()) &
            (~df_data["Country/Region"].str.upper().isin(excluded)) &
            (df_data["Country/Region"].str.upper().isin([
                "ARGENTINA", "BOLIVIA", "CHILE", "COLOMBIA", "COSTA RICA", "CUBA", "DOMINICAN REPUBLIC",
                "ECUADOR", "EL SALVADOR", "GUATEMALA", "HONDURAS", "JAMAICA", "NICARAGUA", "PANAMA",
                "PARAGUAY", "PERU", "PUERTO RICO", "URUGUAY", "VENEZUELA"
            ]))
        ]

        # Render block
        ws_summary.append([""] * 8)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        total = len(latam_df)
        hot = latam_df["IsHot"].sum()
        cold = latam_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Table
        ws_summary.append([])
        cols = ["Item Name", "Country/Region", "Salesperson", "Service", "Stage", "Potential", "Referral Source Category"]
        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in latam_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(ws_summary, start_row=start_row, start_col=1, num_rows=1 + len(latam_df), num_cols=len(cols), align="left")

        # ----------------------------------------
        # 🟩 Market Segment: Spain Desk
        # ----------------------------------------

        market_segment_name = "Spain Desk"

        spain_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "SPAIN")
        ]

        # Render
        ws_summary.append([""] * 8)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        total = len(spain_df)
        hot = spain_df["IsHot"].sum()
        cold = spain_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        ws_summary.append([])
        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in spain_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(ws_summary, start_row=start_row, start_col=1, num_rows=1 + len(spain_df), num_cols=len(cols), align="left")

        # ----------------------------------------
        # 🟩 Market Segment: UK Desk
        # ----------------------------------------

        market_segment_name = "UK Desk"

        uk_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "UNITED KINGDOM")
        ]

        # Spacer row
        ws_summary.append([""] * 8)

        # Section title row (merged)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary
        total = len(uk_df)
        hot = uk_df["IsHot"].sum()
        cold = uk_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in uk_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(uk_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: USA Desk
        # ----------------------------------------

        market_segment_name = "USA Desk"

        usa_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "UNITED STATES")
        ]

        # Spacer
        ws_summary.append([""] * 8)

        # Title row
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary
        total = len(usa_df)
        hot = usa_df["IsHot"].sum()
        cold = usa_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Table
        ws_summary.append([])

        cols = ["Item Name", "Country/Region", "Salesperson", "Service", "Stage", "Potential", "Referral Source Category"]
        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in usa_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(usa_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: Middle East Desk
        # ----------------------------------------

        market_segment_name = "Middle East Desk"

        middle_east_countries = [
            "UAE", "UNITED ARAB EMIRATES", "SAUDI ARABIA", "QATAR", "ISRAEL",
            "KUWAIT", "OMAN", "BAHRAIN", "LEBANON", "JORDAN", "IRAQ", "IRAN"
        ]

        me_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper().isin(middle_east_countries))
        ]

        # Spacer row
        ws_summary.append([""] * 8)

        # Section title row (merged)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary
        total = len(me_df)
        hot = me_df["IsHot"].sum()
        cold = me_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in me_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(me_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: India Desk
        # ----------------------------------------

        market_segment_name = "India Desk"

        india_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "INDIA")
        ]

        # Spacer row
        ws_summary.append([""] * 8)

        # Section title row (merged)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary
        total = len(india_df)
        hot = india_df["IsHot"].sum()
        cold = india_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in india_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(india_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: Euro Desk
        # ----------------------------------------

        market_segment_name = "Euro Desk"

        # Define European countries excluding UK and Spain
        european_countries = [
            "GERMANY", "FRANCE", "ITALY", "NETHERLANDS", "BELGIUM", "SWEDEN", "NORWAY",
            "DENMARK", "FINLAND", "POLAND", "PORTUGAL", "GREECE", "AUSTRIA",
            "CZECH REPUBLIC", "HUNGARY", "IRELAND", "SWITZERLAND", "ROMANIA", "SLOVAKIA"
        ]

        excluded_countries = ["UNITED KINGDOM", "SPAIN"]

        euro_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper().isin(european_countries)) &
            (~df_data["Country/Region"].str.upper().isin(excluded_countries))
        ]

        # Spacer row
        ws_summary.append([""] * 8)

        # Section title row (merged)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary
        total = len(euro_df)
        hot = euro_df["IsHot"].sum()
        cold = euro_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in euro_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(euro_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 🟩 Market Segment: Australia Desk
        # ----------------------------------------

        market_segment_name = "Australia Desk"

        australia_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Country/Region"].fillna("").str.strip().str.upper() == "AUSTRALIA")
        ]

        # Spacer row
        ws_summary.append([""] * 8)

        # Section title row (merged)
        desk_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=desk_title_row, start_column=1, end_row=desk_title_row, end_column=8)

        cell = ws_summary.cell(row=desk_title_row, column=1)
        cell.value = market_segment_name
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Summary
        total = len(australia_df)
        hot = australia_df["IsHot"].sum()
        cold = australia_df["IsCold"].sum()

        ws_summary.append([""] * 8)
        summary_row = ws_summary.max_row + 1
        summary_text = f"There are total {total} active enquiries for {market_segment_name}, out of which {hot} are hot and {cold} are cold."
        ws_summary.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)

        cell = ws_summary.cell(row=summary_row, column=1)
        cell.value = summary_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

        # Enquiry table
        ws_summary.append([])

        cols = [
            "Item Name", "Country/Region", "Salesperson",
            "Service", "Stage", "Potential", "Referral Source Category"
        ]

        ws_summary.append(cols)
        start_row = ws_summary.max_row

        for _, row in australia_df[cols].iterrows():
            ws_summary.append(list(row))

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=1 + len(australia_df),
            num_cols=len(cols),
            align="left"
        )

        # ----------------------------------------
        # 📊 Section: Breakdown by Salesperson (styled section header)
        # ----------------------------------------

        # Add spacer row before the section
        ws_summary.append([""] * 8)
        ws_summary.append([""] * 8)

        # Define title row and merge range
        sales_title_row = ws_summary.max_row + 1
        ws_summary.merge_cells(start_row=sales_title_row, start_column=1, end_row=sales_title_row, end_column=8)

        # Set title cell
        cell = ws_summary.cell(row=sales_title_row, column=1)
        cell.value = "Breakdown by Salesperson"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Add spacer row after section title
        ws_summary.append([""] * 8)

        # ----------------------------------------
        # 📋 Summary Table: Breakdown by Salesperson
        # ----------------------------------------

        # Filter active enquiries with sales info
        sales_df = df_data[
            (df_data["IsActiveNow"] == 1) &
            (df_data["Salesperson"].notna())
        ].copy()

        # Sort by Salesperson for grouping
        sales_df.sort_values(by=["Salesperson"], inplace=True)

        # Table columns to show
        cols = ["Item Name", "Dept", "Service", "Potential"]

        # Spacer before table
        ws_summary.append([])

        # Add table headers
        header_row = ws_summary.max_row + 1
        ws_summary.append(["Salesperson", *cols])
        start_row = ws_summary.max_row  # start of data after header

        # Track rows written
        for salesperson, group in sales_df.groupby("Salesperson"):
            # Add a subtotal row for the salesperson
            ws_summary.append([
                f"{salesperson} (Total: {len(group)})", "", "", "", ""
            ])
            
            # Append each row of enquiries under that salesperson
            for _, row in group.iterrows():
                ws_summary.append([
                    "",  # blank for salesperson column
                    row["Item Name"],
                    row["Dept"],
                    row["Service"],
                    row["Potential"]
                ])

        # Apply formatting
        num_rows = ws_summary.max_row - start_row + 1

        format_table(
            ws_summary,
            start_row=start_row,
            start_col=1,
            num_rows=num_rows,
            num_cols=1 + len(cols),
            align="left"
        )

        

        fixed_width = 20
        for col in range(1, ws_summary.max_column + 1):
            col_letter = get_column_letter(col)
            ws_summary.column_dimensions[col_letter].width = fixed_width

       
        # Apply wrap text to all cells in the Summary Report
        for row in ws_summary.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal if cell.alignment else "left",
                        vertical=cell.alignment.vertical if cell.alignment else "top",
                        wrap_text=True
                    )

    excel_buffer.seek(0)
    return excel_buffer








